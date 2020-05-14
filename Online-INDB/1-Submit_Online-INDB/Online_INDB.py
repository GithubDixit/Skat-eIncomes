import driver as driver
import Login_Add_SENR
import re
import time
import openpyxl

from Login_Add_SENR import webdriver
from Login_Add_SENR import Keys
from Login_Add_SENR import WebDriverWait
from Login_Add_SENR import By
from Login_Add_SENR import Select
from Login_Add_SENR import driver
from Login_Add_SENR import webdriver
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC, wait
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import *
from Login_Add_SENR import USER_INP_SENR

from xpath_repository import *
class indbcreation:
    def __init__(self):
        pass

    # ************ Method to Handle Current Window **************#
    def handle_current_window_method(self):
        handles = driver.window_handles
        for handle in handles:
            driver.switch_to.window(handle)
            # print(driver.title)
            driver.maximize_window()

    # ************ Method to Create Online INDB **************#
    def checkforlawtype(self):
        time.sleep(10)
        if driver.find_element_by_name(scr250_checkbox).is_displayed():
            print("check box visible")
            indbcreation.entercpr(self)
        elif driver.find_element_by_xpath(scr250_eIncomeRadioButton).is_displayed():
            print("E radio button")
            driver.find_element_by_xpath(scr250_eIncomeRadioButton).click()
            indbcreation.entercpr(self)
        else:
            print("Nothing")
            indbcreation.entercpr(self)

    def entercpr(self):
        time.sleep(2)
        driver.find_element_by_xpath(scr250_cprNumber).send_keys("1909580060")
        #driver.find_element_by_xpath('//*[@id="medarbnr"]').send_keys('151072863')
        time.sleep(2)
        # driver.find_element_by_xpath("/html/body/table[2]/tbody/tr[1]/td/table/tbody/tr/td/table/tbody/tr/td[1]/div/table/tbody/tr[8]/td[2]/input").send_keys("1909580060") #Provide CPR-no
        driver.find_element_by_xpath(default_okButton).click()
        time.sleep(3)
        Title = driver.title
        print("Title of new page is", Title)
        if Title == "SKAT, eIndkomst, Personstamoplysninger":
            indbcreation.Personstamoplysninger(self)
        else:
            indbcreation.cpruploadwithpincode(self)

    def Personstamoplysninger(self):
        time.sleep(2)
        if(driver.find_element_by_name(Personstamoplysninger_ansaettelsesDato)).is_displayed():
            driver.find_element_by_name(Personstamoplysninger_ansaettelsesDato).send_keys(20200401)
            driver.find_element_by_name(Personstamoplysninger_fratraedelsesDato).send_keys(20200401)
            driver.find_element_by_name(Personstamoplysninger_skattekortTypeAnvendFra).send_keys(20200401)
            driver.find_element_by_name(Personstamoplysninger_supplerendeMedarbejderNr).send_keys(20200401)
        else:
            driver.find_element_by_xpath(default_okButton).click()
            indbcreation.cpruploadwithpincode(self)

        time.sleep(3)

    def cpruploadwithpincode(self):
        time.sleep(2)
        driver.find_element_by_id("loenPeriodeStartDato").send_keys("2020.09.01")
        time.sleep(3)
        driver.find_element_by_id("loenPeriodeSlutDato").send_keys("2020.09.30")
        time.sleep(3)
        driver.find_element_by_id("dispositionsdato").send_keys("2020.09.15")
        time.sleep(3)
        driver.find_element_by_xpath(scr255_Field13).click()

        # Enter value in Fields
        time.sleep(3)
        # **************** Field 13
        driver.find_element_by_xpath(scr255_Field13).clear()
        driver.find_element_by_xpath(scr255_Field13).send_keys("1000")
        # **************** Field 14
        driver.find_element_by_xpath(scr255_Field14).clear()
        driver.find_element_by_xpath(scr255_Field14).send_keys("5000")
        driver.find_element_by_xpath(okButton).click()
        time.sleep(3)
        indbcreation.checkForDyRulePincode(self)
        time.sleep(3)

    def cpruploadwithoutpincode(self):
        time.sleep(3)
        driver.find_element_by_id("loenPeriodeStartDato").send_keys("2020.09.01")
        time.sleep(3)
        driver.find_element_by_id("loenPeriodeSlutDato").send_keys("2020.09.30")
        time.sleep(3)
        driver.find_element_by_id("dispositionsdato").send_keys("2020.09.15")
        time.sleep(3)
        driver.find_element_by_xpath(scr255_Field13).click()

        # Enter value in Fields
        time.sleep(3)
        # ***** Field 13
        driver.find_element_by_xpath(scr255_Field13).clear()
        driver.find_element_by_xpath(scr255_Field13).send_keys("1000")

        # ***** Field 14
        driver.find_element_by_xpath(scr255_Field14).clear()
        driver.find_element_by_xpath(scr255_Field14).send_keys("5000")

        driver.find_element_by_xpath(okButton).click()
        time.sleep(3)

    def checkForDyRulePincode(self):
        try:
            driver.find_element_by_xpath(scr255_dyrule_pincodepopup).is_displayed()
            print("Dynamic Rule displayed")
            indbcreation.pincodedyvalidation(self)
        except NoSuchElementException:
            print("No Dynamic Rule displayed")
            indbcreation.copy_data_excel_method(self)

    def pincodedyvalidation(self):
        try:
            # Pinkode Popup
            driver.find_element_by_xpath(scr255_dyrule_pincodepopup).is_displayed()
            Dynamic_Rule_Text = driver.find_element_by_xpath(scr255_dyrule_textmessage).text
            print("Pincode Dynamic Validation Rule Error text message:", Dynamic_Rule_Text)

            Dynamic_Rule = re.search('\(([^)]+)', Dynamic_Rule_Text).group(1)  # Pick rule Name
            print("The Error is due to Dynamic Rule Number :", Dynamic_Rule)
            driver.find_element_by_xpath(scr255_dyrule_popup_tiblage).click()  # click Tilbage

            indbcreation.dynamicScreenNavigation(self)

            # click Dynamiske valideringsregler menu
            driver.find_element_by_xpath(Dynamiske_valideringsregler_link).click()
            # Move to current new window

            indbcreation.handle_current_window_method(self)
            time.sleep(5)

            # Enter Fejlnr in the Dynamic Rule Search field
            driver.find_element_by_xpath(search_dyrulename_textbox).send_keys(Dynamic_Rule)
            driver.find_element_by_xpath(dyrule_search_button).click()  # Click Sog
            time.sleep(3)

            # Select the Rule
            driver.find_element_by_xpath(dyrule_open).click()
            time.sleep(5)

            # Copy value of Pincode
            Pinkode = driver.find_element_by_xpath(dyrule_pincode).get_attribute("value")
            print(Pinkode)

            # Click on Indkomst menu
            driver.find_element_by_xpath(Indkomst_menu).click()
            indbcreation.handle_current_window_method(self)

            # click Afslut button
            driver.find_element_by_xpath(Afslut_button).click()

            # Click on Indberet lønoplysninger - online to create INDB
            indbcreation.handle_current_window_method(self)
            driver.find_element_by_xpath(Indberet_lønoplysninger_link).click()
            indbcreation.handle_current_window_method(self)
            time.sleep(3)
            indbcreation.cpruploadwithoutpincode(self)
            # Enter Pincode and click ok
            driver.find_element_by_xpath(dyrule_pincode).send_keys(Pinkode)
            driver.find_element_by_xpath(scr255_dyrule_popup_ok).click()
        except NoSuchElementException:
            print(' ')

        try:
            if driver.find_element_by_xpath(scr260_Error_advice_check).is_displayed():
                Fejl_Advis = driver.find_element_by_xpath(scr260_Error_advice_check).text
            if Fejl_Advis == 'Fejl':
                print("Error found proceeding to submit INDB")
            elif Fejl_Advis == 'Advis':
                print("Only Adivce found proceeding to submit INDB")
                driver.find_element_by_xpath(okButton).click()
                time.sleep(5)
                indbcreation.copy_data_excel_method(self)  # Call Method to
                # Click Afslut Button to close INDB Creation Screen
                driver.find_element_by_xpath(scr261_260_Afsult_button).click()
            else:
                print("No Adivce/Error found")
                driver.find_element_by_xpath(okButton).click()
                time.sleep(5)
                indbcreation.copy_data_excel_method(self)  # Call Method to
                # Click Afslut Button to close INDB Creation Screen
                driver.find_element_by_xpath(scr261_260_Afsult_button).click()
        except NoSuchElementException:
            time.sleep(5)

    def dynamicScreenNavigation(self):
        driver.find_element_by_xpath(Indkomst_menu).click()  # Click Indkomst menu
        time.sleep(3)
        # Click Systemadministration menu
        indbcreation.handle_current_window_method(self)
        driver.find_element_by_xpath(System_admin_link).click()
        indbcreation.handle_current_window_method(self)
        time.sleep(5)

    # ************ Method to Copy Data in Excel **************#
    def copy_data_excel_method(self):
        # ************** Load Excel **************#
        path = (
            r"C:\Users\AbhinavDixit\PycharmProjects\Skat-eIncomes\Online-INDB\1-Submit_Online-INDB\Online_INDB_Excel.xlsx")
        workbook = openpyxl.load_workbook(path)  # Load Workbook
        sheet = workbook['INDBID']
        time.sleep(5)
        INDB_ID = driver.find_element_by_xpath(copy_INDBID).text
        print(INDB_ID)
        # Copy INDB_ID in to Excel
        i = 2
        while sheet.cell(row=i, column=1).value != None:
            i = i + 1
        sheet.cell(i, 1).value = INDB_ID
        workbook.save(path)
        driver.find_element_by_xpath(okButton).click()
        driver.quit()

    # Click to create Online INDB

time.sleep(15)
driver.find_element_by_xpath(Indberet_lønoplysninger_link).click()
time.sleep(10)
A1 = indbcreation()
A1.handle_current_window_method()
A1.checkforlawtype()
