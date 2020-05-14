from selenium.common.exceptions import TimeoutException

import Login_Add_SENR

from Login_Add_SENR import webdriver
from Login_Add_SENR import Keys
from Login_Add_SENR import WebDriverWait
from Login_Add_SENR import By
from Login_Add_SENR import Select
from Login_Add_SENR import driver
from Login_Add_SENR import webdriver

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
import time
import openpyxl
global NullINDB_ID
from xpath_repository import *

class nullindb():

    def handle_current_window_method(self):
        handles = driver.window_handles
        for handle in handles:
            driver.switch_to.window(handle)
        # print(driver.title)
            driver.maximize_window()

    def nullindbmethod(self):
        time.sleep(3)
        driver.find_element_by_css_selector(nullINDB_link).click()
        nullindb.handle_current_window_method(self)
        time.sleep(5)
        try:
            WebDriverWait(driver, 20 ).until(EC.presence_of_element_located((By.XPATH,'generelIndberetning.indberetningsident')))
            print("Page Loaded")
            time.sleep(6)
            driver.find_element_by_name("generelIndberetning.indberetningsident").click()
            # driver.find_element_by_xpath(            "/html/body/table[2]/tbody/tr[1]/td/table/tbody/tr/td/table/tbody/tr/td[1]/div/table/tbody/tr[8]/td[2]/select")
            driver.find_element_by_xpath(
                "/html/body/table[2]/tbody/tr[1]/td/table/tbody/tr/td/table/tbody/tr/td[1]/div/table/tbody/tr[8]/td[2]/select").send_keys(
                "Bagud")
            time.sleep(3)
            driver.find_element_by_xpath("//*[@id='defaultButton']").click()

            NullINDB_ID = driver.find_element_by_xpath(
                "/html/body/table[2]/tbody/tr[1]/td/table/tbody/tr/td/table/tbody/tr/td[1]/div/table[2]/tbody/tr[1]/td[2]").text
              # Copy INDB_ID in to Excel
            path = (
                r"C:\Users\AbhinavDixit\PycharmProjects\Skat-eIncomes\Online-INDB\1-Submit_Online-INDB\Online_INDB_Excel.xlsx")
            workbook = openpyxl.load_workbook(path)  # Load Workbook
            NullINDB_Sheet = workbook['NullINDB']  # Load active sheet and save in sheet object
            # Copy INDB_ID in to Excel
            i = 1
            while NullINDB_Sheet.cell(row=i, column=1).value != None:
                i = i + 1
            NullINDB_Sheet.cell(i, 1).value = NullINDB_ID
            workbook.save(path)
            driver.find_element_by_xpath(
                "/html/body/table[2]/tbody/tr[2]/td/table/tbody/tr/td[1]/input").click()  # Save NullINDB
            # driver.find_element_by_xpath("/html/body/table[3]/tbody/tr/td[2]/input").click()  # Close window
        except TimeoutException:
            print("Loading took too much time!")

        #Title = driver.title
        #print("Title of new page is", Title)


A1 = nullindb()
A1.nullindbmethod()
