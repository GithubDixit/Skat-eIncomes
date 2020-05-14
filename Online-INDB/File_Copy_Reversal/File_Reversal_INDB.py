# Call Login python file
import driver as driver
import Login_Add_SENR
from Login_Add_SENR import webdriver
from Login_Add_SENR import Keys
from Login_Add_SENR import WebDriverWait
from Login_Add_SENR import By
from Login_Add_SENR import Select
from Login_Add_SENR import driver
from Login_Add_SENR import webdriver

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC, wait
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import *

import time
import xlrd  # import package to read data from Excel
import openpyxl
from openpyxl import load_workbook



class indbreversal():
    global recent_indbid, file_path
    # ************ Method to Handle Current Window **************#
    def handle_current_window_method(self):
        handles = driver.window_handles
        for handle in handles:
            driver.switch_to.window(handle)
            # print(driver.title)
            driver.maximize_window()

    def reversalScreenNavigation(self):
        # Click on Forespørg/Kopiér/Tilbagefør indberetninger
        time.sleep(5)
        driver.find_element_by_xpath(
            "/html/body/table[2]/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td[1]/table/tbody/tr[1]/td/a[7]").click()

        # Move to current new window
        indbreversal.handle_current_window_method(self)
        time.sleep(5)

        # select search option Indberetnings-ID from drop down
        driver.find_element_by_xpath(
            "/html/body/table[2]/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr[7]/td[2]/select").is_displayed()
        driver.find_element_by_xpath(
            "/html/body/table[2]/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr[7]/td[2]/select").send_keys(
            "IndberetningsID")
        time.sleep(3)
        indbreversal.readINDBfromexcel(self)


    def readINDBfromexcel(self):
        # Pass entire file path as parameter
        # set file path
        file_path = (
            r"C:\Users\AbhinavDixit\PycharmProjects\Skat-eIncomes\Online-INDB\1-Submit_Online-INDB\Online_INDB_Excel.xlsx")
        book = xlrd.open_workbook(file_path)
        aa = book.sheet_by_index(0)
        list_ad = []
        total_rows = aa.nrows
        print("Total number of Rows are:", total_rows)
        for i in range(1, total_rows):
            list_ad.append(aa.cell(rowx=i, colx=0).value)
            print("Length of List", len(list_ad))
            print("LIST", list_ad)
        print("THE FINAL LIST IS ", list_ad)
        time.sleep(3)
        recent_indbid = list_ad[-1]
        # Enter INDB ID to be searched
        driver.find_element_by_xpath(
            "/html/body/table[2]/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr[7]/td[2]/input").send_keys(
            recent_indbid)
        driver.find_element_by_xpath("//*[@id='defaultButton']").click() # Press Sog Button
        indbreversal.indbreversalmethod(self,recent_indbid)

    def indbreversalmethod(self,recent_indbid):

        try:
            WebDriverWait(driver, 3).until(EC.alert_is_present(),
                                           'Timed out waiting for PA creation ' +
                                           'confirmation popup to appear.')

            alert = driver.switch_to.alert
            print(alert.text)
            time.sleep(5)
            alert.accept()
            print("Popup Alert Message Appeared and Accepted while searching INDB")
        except TimeoutException:
            print("No Popup Alert Message Appeared while searching INDB")
            time.sleep(5)

        try:
            INDB_ID = driver.find_element_by_xpath(
                "/html/body/table[2]/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr[18]/td/table/tbody/tr[2]/td[3]").text
            ART_VALUE = driver.find_element_by_xpath(
                "/html/body/table[2]/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr[18]/td/table/tbody/tr[2]/td[7]").text
            TILBAGEFORT_VALUE = driver.find_element_by_xpath(
                "/html/body/table[2]/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr[18]/td/table/tbody/tr[2]/td[8]").text
            print("INDB ID is :", INDB_ID)
            print("ART Value is:", ART_VALUE)
            print("TILBAGEFORT VALUE is:", TILBAGEFORT_VALUE)
            print(recent_indbid)
            if driver.find_element_by_xpath(
                    "/html/body/table[2]/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr[18]/td/table/tbody/tr[2]/td[3]").is_displayed():
                if INDB_ID == recent_indbid and ART_VALUE == 'I' and TILBAGEFORT_VALUE == 'Nej':
                    driver.find_element_by_xpath(
                        "/html/body/table[2]/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr[18]/td/table/tbody/tr[2]/td[11]/input").click()  # select Radio Button
                    driver.find_element_by_xpath(
                        "/html/body/table[2]/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr[16]/td[2]/input").click()  # click Tilbagefør button
                    time.sleep(3)

                    New_INDBREV_ID = driver.find_element_by_xpath(
                        "/html/body/table[2]/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr[7]/td[2]/input").is_displayed()
                    print(New_INDBREV_ID)  # Check whether New INDB REV ID field is dispalyed

                    New_INDB_REV_ID = driver.find_element_by_xpath(
                        "/html/body/table[2]/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr[7]/td[2]/input").get_attribute(
                        "value")
                    print("New_INDB_REV_ID is:-", New_INDB_REV_ID)  # print New INDB REV ID

                    Hovedindberetningsident = driver.find_element_by_xpath(
                        "/html/body/table[2]/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr[4]/td[2]/input").get_attribute(
                        "value")
                    print("Hovedindberetningsident is:", Hovedindberetningsident)

                    # Copy INDB ID to Excel File
                    time.sleep(5)
                    workbook = openpyxl.load_workbook(file_path)  # Load Workbook
                    sheet = workbook['REV_INDBID']
                    print(New_INDBREV_ID)
                    # Copy INDB_ID in to Excel
                    i = 2
                    while sheet.cell(row=i, column=1).value != None:
                        i = i + 1
                    sheet.cell(i, 1).value = INDB_ID
                    sheet.cell(i, 2).value = New_INDB_REV_ID
                    sheet.cell(i, 3).value = Hovedindberetningsident
                    workbook.save(file_path)
                else:
                    print('Since TILBAGEFORT VALUE is: Ja File cannot be reversed ')
        except NoSuchElementException:
                print('  ')

        # Click bekraft button to commit Tilbage (Reversal)
        driver.find_element_by_xpath("//*[@id='defaultButton']").click()
        try:
            WebDriverWait(driver, 3).until(EC.alert_is_present(),
                                       'Timed out waiting for PA creation ' +
                                       'confirmation popup to appear.')

            alert = driver.switch_to.alert
            print(alert.text)
            time.sleep(5)
            alert.accept()
            print("alert accepted")
        except TimeoutException:
            print("no popup alert")
            time.sleep(5)

A1 = indbreversal()
A1.reversalScreenNavigation()