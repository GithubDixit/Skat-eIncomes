import driver as driver
import Login_Add_SENR
import re
import time
import openpyxl

from Login_Add_SENR import webdriver
from Login_Add_SENR import Keys
from Login_Add_SENR import WebDriverWait
from Login_Add_SENR import By
from Login_Add_SENR import Select
from Login_Add_SENR import driver
from Login_Add_SENR import webdriver
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC, wait
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import *

class test:
    def __init__(self):
        pass
    # ************ Method to Handle Current Window **************#
    def handle_current_window_method(self):
        handles = driver.window_handles
        for handle in handles:
            driver.switch_to.window(handle)
            # print(driver.title)
            driver.maximize_window()

    # ************ Method to Create Online INDB **************#
    def checkforlawtype(self):
        try:
            time.sleep(5)
            driver.find_elements_by_xpath("/html/body/table[2]/tbody/tr[1]/td/table/tbody/tr/td/table/tbody/tr/td[1]/div/table/tbody/tr[5]/td[2]/span[2]")
            print("Visible")
            test.selecteIndkomst(self)
        except NoSuchElementException:
            print("not visible")
            test.cpruploadwithpincode(self)

    def selecteIndkomst(self):
        # Select EIncome Radio Button
        time.sleep(3)
        driver.find_element_by_xpath(
            "/html/body/table[2]/tbody/tr[1]/td/table/tbody/tr/td/table/tbody/tr/td[1]/div/table/tbody/tr[4]/td[2]/span[2]/input").click()
        test.cpruploadwithpincode(self)

    def cpruploadwithpincode(self):
        time.sleep(2)
        #driver.find_element_by_id("cprnr").send_keys("1909580060") or
        #driver.find_element_by_xpath("//*[@id='cprnr']").send_keys("1909580060")
        driver.find_element_by_xpath("/html/body/table[2]/tbody/tr[1]/td/table/tbody/tr/td/table/tbody/tr/td[1]/div/table/tbody/tr[8]/td[2]/input").send_keys("1909580060")
        time.sleep(2)
            #driver.find_element_by_xpath("/html/body/table[2]/tbody/tr[1]/td/table/tbody/tr/td/table/tbody/tr/td[1]/div/table/tbody/tr[8]/td[2]/input").send_keys("1909580060") #Provide CPR-no
        driver.find_element_by_xpath("//*[@id='defaultButton']").click()
        time.sleep(3)
        driver.find_element_by_id("loenPeriodeStartDato").send_keys("2020.09.01")
        time.sleep(3)
        driver.find_element_by_id("loenPeriodeSlutDato").send_keys("2020.09.30")
        time.sleep(3)
        driver.find_element_by_id("dispositionsdato").send_keys("2020.09.15")
        time.sleep(3)
        driver.find_element_by_xpath(
            "/html/body/table[2]/tbody/tr[1]/td/table/tbody/tr/td/table/tbody/tr/td[1]/div/table/tbody/tr[15]/td[2]/input").click()

        # Enter value in Fields
        time.sleep(3)
        # **************** Field 13
        driver.find_element_by_xpath(
            "/html/body/table[2]/tbody/tr[1]/td/table/tbody/tr/td/table/tbody/tr/td[1]/div/table/tbody/tr[15]/td[2]/input").clear()
        driver.find_element_by_xpath(
            "/html/body/table[2]/tbody/tr[1]/td/table/tbody/tr/td/table/tbody/tr/td[1]/div/table/tbody/tr[15]/td[2]/input").send_keys(
            "1000")
        # **************** Field 14
        driver.find_element_by_xpath(
            "/html/body/table[2]/tbody/tr[1]/td/table/tbody/tr/td/table/tbody/tr/td[1]/div/table/tbody/tr[15]/td[4]/input").clear()
        driver.find_element_by_xpath("/html/body/table[2]/tbody/tr[1]/td/table/tbody/tr/td/table/tbody/tr/td[1]/div/table/tbody/tr[15]/td[4]/input").send_keys("5000")
        driver.find_element_by_xpath("/html/body/table[2]/tbody/tr[2]/td/table/tbody/tr/td[2]/input").click()
        time.sleep(3)
        test.checkForDyRulePincode(self)
        time.sleep(3)

    def cpruploadwithoutpincode(self):
        time.sleep(2)
        #driver.find_element_by_id("cprnr").send_keys("1909580060") or
        #driver.find_element_by_xpath("//*[@id='cprnr']").send_keys("1909580060")
        driver.find_element_by_xpath("/html/body/table[2]/tbody/tr[1]/td/table/tbody/tr/td/table/tbody/tr/td[1]/div/table/tbody/tr[8]/td[2]/input").send_keys("1909580060")
        time.sleep(2)
            #driver.find_element_by_xpath("/html/body/table[2]/tbody/tr[1]/td/table/tbody/tr/td/table/tbody/tr/td[1]/div/table/tbody/tr[8]/td[2]/input").send_keys("1909580060") #Provide CPR-no
        driver.find_element_by_xpath("//*[@id='defaultButton']").click()
        time.sleep(3)
        driver.find_element_by_id("loenPeriodeStartDato").send_keys("2020.09.01")
        time.sleep(3)
        driver.find_element_by_id("loenPeriodeSlutDato").send_keys("2020.09.30")
        time.sleep(3)
        driver.find_element_by_id("dispositionsdato").send_keys("2020.09.15")
        time.sleep(3)
        driver.find_element_by_xpath(
            "/html/body/table[2]/tbody/tr[1]/td/table/tbody/tr/td/table/tbody/tr/td[1]/div/table/tbody/tr[15]/td[2]/input").click()

        # Enter value in Fields
        time.sleep(3)
        # **************** Field 13
        driver.find_element_by_xpath(
            "/html/body/table[2]/tbody/tr[1]/td/table/tbody/tr/td/table/tbody/tr/td[1]/div/table/tbody/tr[15]/td[2]/input").clear()
        driver.find_element_by_xpath(
            "/html/body/table[2]/tbody/tr[1]/td/table/tbody/tr/td/table/tbody/tr/td[1]/div/table/tbody/tr[15]/td[2]/input").send_keys(
            "1000")
        # **************** Field 14
        driver.find_element_by_xpath(
            "/html/body/table[2]/tbody/tr[1]/td/table/tbody/tr/td/table/tbody/tr/td[1]/div/table/tbody/tr[15]/td[4]/input").clear()
        driver.find_element_by_xpath("/html/body/table[2]/tbody/tr[1]/td/table/tbody/tr/td/table/tbody/tr/td[1]/div/table/tbody/tr[15]/td[4]/input").send_keys("5000")
        driver.find_element_by_xpath("/html/body/table[2]/tbody/tr[2]/td/table/tbody/tr/td[2]/input").click()
        time.sleep(3)


    def checkForDyRulePincode(self):
        try:
            driver.find_element_by_xpath(
                "/html/body/table[2]/tbody/tr[1]/td/table/tbody/tr/td/table/tbody/tr/td[1]/div/div/div/div[2]/input[1]").is_displayed()
            print("Dynamic Rule displayed")
            test.pincodedyvalidation(self)
        except NoSuchElementException:
            print("No Dynamic Rule displayed")
            test.copy_data_excel_method(self)

    def pincodedyvalidation(self):
        try:
            #Pinkode Popup
            driver.find_element_by_xpath(
            "/html/body/table[2]/tbody/tr[1]/td/table/tbody/tr/td/table/tbody/tr/td[1]/div/div/div/div[2]/input[1]").is_displayed()
            Dynamic_Rule_Text = driver.find_element_by_xpath(
            "/html/body/table[2]/tbody/tr[1]/td/table/tbody/tr/td/table/tbody/tr/td[1]/div/div/div/div[1]/div[1]").text
            print("Pincode Dynamic Validation Rule Error text message:", Dynamic_Rule_Text)
            Dynamic_Rule = re.search('\(([^)]+)', Dynamic_Rule_Text).group(1)  # Pick rule Name
            print("The Error is due to Dynamic Rule Number :", Dynamic_Rule)
            driver.find_element_by_xpath(
            "/html/body/table[2]/tbody/tr[1]/td/table/tbody/tr/td/table/tbody/tr/td[1]/div/div/div/div[2]/input[1]").click()  # click Tilbage
            test.dynamicScreenNavigation(self)
        # click Dynamiske valideringsregler menu
            driver.find_element_by_xpath(
            "/html/body/table[2]/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr[2]/td/a[13]").click()

        # Move to current new window
            test.handle_current_window_method(self)
            time.sleep(5)

        # Enter Fejlnr in the Dynamic Rule Search field
            driver.find_element_by_xpath(
            "/html/body/table[2]/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr/td/table[1]/tbody/tr[3]/td[2]/input").send_keys(
            Dynamic_Rule)
            driver.find_element_by_xpath(
            "/html/body/table[2]/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr/td/table[1]/tbody/tr[4]/td[6]/input").click()  # Click Sog
            time.sleep(3)

        # Select the Rule
            driver.find_element_by_xpath(
            "/html/body/table[2]/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr/td/table[2]/tbody/tr[2]/td[3]").click()
            time.sleep(5)

        # Copy value of Pincode
            Pinkode = driver.find_element_by_xpath("//*[@id='pinkode']").get_attribute("value")
            print(Pinkode)

        # Click on Indkomst menu
            driver.find_element_by_xpath("/html/body/div[2]/span[1]/a").click()
            test.handle_current_window_method(self)
        # click Afslut button
            driver.find_element_by_xpath(
            "/html/body/table[2]/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr[2]/td/input").click()

        # Click on Indberet lønoplysninger - online to create INDB
            test.handle_current_window_method(self)
            driver.find_element_by_xpath(
            "/html/body/table[2]/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td[1]/table/tbody/tr[1]/td/a[3]").click()
            test.handle_current_window_method(self)
            time.sleep(3)
            test.cpruploadwithoutpincode(self)
        # Enter Pincode and click ok
            driver.find_element_by_xpath("//*[@id='pinkode']").send_keys(Pinkode)
            driver.find_element_by_xpath(
            "/html/body/table[2]/tbody/tr[1]/td/table/tbody/tr/td/table/tbody/tr/td[1]/div/div/div/div[2]/input[2]").click()
        except NoSuchElementException:
            print(' ')

        try:
            if driver.find_element_by_xpath(
                "/html/body/table[2]/tbody/tr[1]/td/table/tbody/tr/td/table/tbody/tr/td[1]/div/table[2]/tbody/tr[2]/td[3]").is_displayed():
                Fejl_Advis = driver.find_element_by_xpath(
                "/html/body/table[2]/tbody/tr[1]/td/table/tbody/tr/td/table/tbody/tr/td[1]/div/table[2]/tbody/tr[2]/td[3]").text
            if Fejl_Advis == 'Fejl':
                print("Error found proceeding to submit INDB")
            elif Fejl_Advis == 'Advis':
                print("Only Adivce found proceeding to submit INDB")
                driver.find_element_by_xpath("/html/body/table[2]/tbody/tr[2]/td/table/tbody/tr/td[2]/input").click()
                time.sleep(5)
                test.copy_data_excel_method(self)  # Call Method to
                # Click Afslut Button to close INDB Creation Screen
                driver.find_element_by_xpath(
                    "/html/body/table[2]/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr/td/table[5]/tbody/tr/td[2]/input[1]").click()
            else:
                print("No Adivce/Error found")
                driver.find_element_by_xpath("/html/body/table[2]/tbody/tr[2]/td/table/tbody/tr/td[2]/input").click()
                time.sleep(5)
                test.copy_data_excel_method(self)  # Call Method to
                # Click Afslut Button to close INDB Creation Screen
                driver.find_element_by_xpath(
                    "/html/body/table[2]/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr/td/table[5]/tbody/tr/td[2]/input[1]").click()
        except NoSuchElementException:
            time.sleep(5)

    def dynamicScreenNavigation(self):
        driver.find_element_by_xpath("/html/body/div[2]/span[1]/a").click()  # Click Indkomst menu
        time.sleep(3)
        # Click Systemadministration menu
        test.handle_current_window_method(self)
        driver.find_element_by_xpath(
        "/html/body/table[2]/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td[1]/table/tbody/tr[1]/td/a[24]").click()
        test.handle_current_window_method(self)
        time.sleep(5)
    # ************ Method to Copy Data in Excel **************#
    def copy_data_excel_method(self):
        # ************** Load Excel **************#
        path = (
            r"C:\Users\AbhinavDixit\PycharmProjects\Skat-eIncomes\Online-INDB\1-Submit_Online-INDB\Online_INDB_Excel.xlsx")
        workbook = openpyxl.load_workbook(path)  # Load Workbook
        sheet = workbook['INDBID']
        time.sleep(5)
        INDB_ID = driver.find_element_by_xpath(
            "/html/body/table[2]/tbody/tr[1]/td/table/tbody/tr/td/table/tbody/tr/td[1]/div/table[2]/tbody/tr[3]/td[2]").text
        print(INDB_ID)
        # Copy INDB_ID in to Excel
        i = 2
        while sheet.cell(row=i, column=1).value != None:
            i = i + 1
        sheet.cell(i, 1).value = INDB_ID
        workbook.save(path)
        driver.find_element_by_xpath("/html/body/table[2]/tbody/tr[2]/td/table/tbody/tr/td[2]/input").click()

    # Click to create Online INDB
driver.find_element_by_xpath(
                '/html/body/table[2]/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td[1]/table/tbody/tr[1]/td/a[3]').click()
A1 = test()
A1.handle_current_window_method()
A1.checkforlawtype()
